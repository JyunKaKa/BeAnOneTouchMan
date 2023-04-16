#備品紀錄單
#


import tkinter as tk
import os
import datetime
import shutil
import time
from tkinter.constants import NO
import win32com.client as win32 
import openpyxl
import pyautogui
#全域變數
v1 = []#儲存EXCEL的檔案，可以去CONSTANT.PY查表
lblString = ""#顯示資料讓使用者知道有沒有成功建立。

#建立視窗
root = tk.Tk()
root.title("備品開單小幫手")

#創建標籤
label1 = tk.Label(root, text="請輸入請購編號:")
label1.grid(row=0, column=0)

#創建文字框
num = tk.StringVar()
entry1 = tk.Entry(root, textvariable=num)
entry1.grid(row=0, column=1)

#創建執行狀況顯示文字框
label2 = tk.Label(root, text="執行結果")
label2.grid(row=2)

from datetime import datetime, timedelta
import locale
locale.setlocale(locale.LC_CTYPE, 'Chinese')
def to_ROC_date_today():
    # 取得今天的日期
    today = datetime.now().date()
    # 計算出民國年
    roc_year = today.year - 1911
    # 格式化輸出字串
    roc_date_str = f'{roc_year}{today.strftime("年%m月%d日")}'
    return str(roc_date_str)

def tax(money):
    s = int(money)
    tax =str(round(s*0.05))
    return addComma(tax)

def moneyWithTax(money):
    s = int(money)
    moneyWithTax =str(round(s*1.05))
    return addComma(moneyWithTax)

def addComma(s):
    #s = s.replace("NTD", "") # 刪除 NTD
    result = ""
    count = 0
    for digit in s[::-1]:
        if count == 3:
            result = "," + result
            count = 0
        result = digit + result
        count += 1

    #result = "NTD " + result
    print(result)
    return result

def depart(str):
    departStr = ""
    if str=="儀控" or str=="機械" or str=="電氣":
        departStr="維修"
    elif str=="管理課":
        departStr="管理"
    elif str=="工安課":
        departStr="工安"
    elif str=="水處理":
        departStr="運轉"
    else:
        msg =tk.Message(root,text="注意XX課有誤",font=("Algerian",18,"bold"),bg='#ADFEDC',fg='#00CACA')
    return departStr

#按下Enter按鈕時執行的功能
def execute():
    global num
    global v1
    global lblString

    #取得num的值
    num_value = num.get()
    
        ##這是將excel檔案複製到txt檔案中


    # 指定 Excel 檔案路徑
    #path = r"D:\Users\E941\Desktop\手工驗收單\112工程發包俊龍版2.xlsx"
    path = r"../../112備品俊龍板_0410.xlsx"


    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(path)

    # 宣告欲尋找的字串
    #num = "20231185"

    # 宣告一個空的 List 變數
    #v1 = []

    # 尋找 num 的位置
    for sheet in wb.worksheets:
        for row in sheet.rows:
            for cell in row:
                if str(cell.value) == num_value:
                    for i in range(26):
                        v1.append(str(row[i].value))
                        print(v1[i])


    if len(v1)==0:
        lblString = "於找不到這一筆資料。\n"
        label2['text'] = lblString
        return
    else:
        for i in range(23):
            if v1[i]==None:
                v1[i]="0"

#根據最後議價金額判定使用哪一張議價單
#要怎麼判定獨家? 再做一個按鈕判定?
    lastCost =int(v1[18]) 
    if lastCost<8000:

        # 複製文件
        #src_file = r"D:\Users\E941\Desktop\採購表單\07_驗收表單\F-AD-2-09-05A維修驗收單.docx"
        
        src_file = r"../../採購表單/01比價會議紀錄表_空白/比（議）價會議紀錄  金額8000 以下.docx"

        #這段程式碼，在未來想要顯示品名作為word檔案名稱時候，要使用，
        new_filename ="_"#v1[3]
        # 檢查檔案名稱是否合法，若不合法則替換特殊符號
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|',' ']
        for char in invalid_chars:
            if char in new_filename:
                new_filename = new_filename.replace(char, '_')
        dst_file_name = num_value + new_filename + "_8k以下議價紀錄"

        #更換目前位置change directory
        # print("目前位置為"+os.getcwd())
        os.chdir("..\\..\\採購表單/")
        dst_folder = str(os.getcwd()) +"\\"
        #print("2目前位置為"+os.getcwd())
            #os.path.join()是把字串黏接起來
        dst_file = os.path.join(dst_folder, dst_file_name + ".docx")
        print("dst_file是"+ dst_file)

        #要將相對位置跳回去
        os.chdir("..\\pyT\\priceRecord")
        #卡在這邊無法copy
        shutil.copy(src_file,dst_file)#這一行srcfile可以相對位置，dst_file只能絕對位置。
        #os.system(f"copy {src_file} {dst_file}")

        # 打開文件
        word = win32.gencache.EnsureDispatch('Word.Application')
        doc = word.Documents.Open(dst_file)

        # 移動游標並進行文字操作
        word.Selection.MoveDown()
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[3])#品名

        word.Selection.Find.Execute("時間：")
        word.Selection.MoveRight()
        today = to_ROC_date_today()#可能要去掉民國
        word.Selection.TypeText(today)
        #for i in range(8):
            #word.Selection.Delete()
        word.Selection.Find.Execute("本案預算NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[17]))#輸入預算
        word.Selection.Find.Execute("廠商名稱")
        word.Selection.MoveRight()
        word.Selection.MoveDown()
        word.Selection.TypeText(v1[16])#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD"+ addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("本案請廠商「")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[16])#輸入廠商
        word.Selection.Find.Execute("最後價格為NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("交貨期限:")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[12])#輸入工作天數
        word.Selection.Find.Execute("會課:")
        word.Selection.MoveLeft()#前往單價
        word.Selection.MoveRight()#前往單價
        word.Selection.TypeText(depart(v1[6]))#輸入XX課

    elif lastCost>=8000 and lastCost<=100000:
        
        # 複製文件
        #src_file = r"D:\Users\E941\Desktop\採購表單\07_驗收表單\F-AD-2-09-05A維修驗收單.docx"
        
        src_file = r"../../採購表單/01比價會議紀錄表_空白/比（議）價會議紀錄  金額8000以上.docx"

        #這段程式碼，在未來想要顯示品名作為word檔案名稱時候，要使用，
        new_filename ="_"#v1[3]
        # 檢查檔案名稱是否合法，若不合法則替換特殊符號
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|',' ']
        for char in invalid_chars:
            if char in new_filename:
                new_filename = new_filename.replace(char, '_')
        dst_file_name = num_value + new_filename + "_議價紀錄"

        #更換目前位置change directory
        # print("目前位置為"+os.getcwd())
        os.chdir("..\\..\\採購表單/")
        dst_folder = str(os.getcwd()) +"\\"
        #print("2目前位置為"+os.getcwd())
            #os.path.join()是把字串黏接起來
        dst_file = os.path.join(dst_folder, dst_file_name + ".docx")
        print("dst_file是"+ dst_file)

        #要將相對位置跳回去
        os.chdir("..\\pyT\\priceRecord")
        #卡在這邊無法copy
        shutil.copy(src_file,dst_file)#這一行srcfile可以相對位置，dst_file只能絕對位置。
        #os.system(f"copy {src_file} {dst_file}")

        # 打開文件
        word = win32.gencache.EnsureDispatch('Word.Application')
        doc = word.Documents.Open(dst_file)

        # 移動游標並進行文字操作
        word.Selection.MoveDown()
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[3])#品名

        word.Selection.Find.Execute("時間：")
        word.Selection.MoveRight()
        today = to_ROC_date_today()#可能要去掉民國
        word.Selection.TypeText(today)
        #for i in range(8):
            #word.Selection.Delete()
        word.Selection.Find.Execute("本案預算NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[17]))#輸入預算
        word.Selection.Find.Execute("廠商名稱")
        word.Selection.MoveRight()
        word.Selection.MoveDown()
        word.Selection.TypeText(v1[16])#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD"+addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("本案請廠商「")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[16])#輸入廠商
        word.Selection.Find.Execute("最後價格為NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("4.	交貨期限:")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[12]+"內交貨。")#輸入工作天數
        word.Selection.Find.Execute("會課:")
        word.Selection.MoveLeft()#前往單價
        word.Selection.MoveRight()#前往單價
        word.Selection.TypeText(depart(v1[6]))#輸入XX課
    elif lastCost>=100000 and lastCost<150000:
            
        # 複製文件
        #src_file = r"D:\Users\E941\Desktop\採購表單\07_驗收表單\F-AD-2-09-05A維修驗收單.docx"
        
        src_file = r"../../採購表單/01比價會議紀錄表_空白/三家以上-- 比（議）價會議紀錄  金額10萬~未達15萬.docx"

        #這段程式碼，在未來想要顯示品名作為word檔案名稱時候，要使用，
        new_filename ="_"#v1[3]
        # 檢查檔案名稱是否合法，若不合法則替換特殊符號
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|',' ']
        for char in invalid_chars:
            if char in new_filename:
                new_filename = new_filename.replace(char, '_')
        dst_file_name = num_value + new_filename + "_10Wto15W議價紀錄"

        #更換目前位置change directory
        # print("目前位置為"+os.getcwd())
        os.chdir("..\\..\\採購表單/")
        dst_folder = str(os.getcwd()) +"\\"
        #print("2目前位置為"+os.getcwd())
            #os.path.join()是把字串黏接起來
        dst_file = os.path.join(dst_folder, dst_file_name + ".docx")
        print("dst_file是"+ dst_file)

        #要將相對位置跳回去
        os.chdir("..\\pyT\\priceRecord")
        #卡在這邊無法copy
        shutil.copy(src_file,dst_file)#這一行srcfile可以相對位置，dst_file只能絕對位置。
        #os.system(f"copy {src_file} {dst_file}")

        # 打開文件
        word = win32.gencache.EnsureDispatch('Word.Application')
        doc = word.Documents.Open(dst_file)

        # 移動游標並進行文字操作
        word.Selection.MoveDown()
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[3])#品名

        word.Selection.Find.Execute("時間：")
        word.Selection.MoveRight()
        today = to_ROC_date_today()
        word.Selection.TypeText(today)
        #for i in range(8):
            #word.Selection.Delete()
        word.Selection.Find.Execute("本案預算NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[17]))#輸入預算
        word.Selection.Find.Execute("廠商名稱")
        word.Selection.MoveRight()
        word.Selection.MoveDown()
        word.Selection.TypeText(v1[16])#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD"+addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("最後「")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[16])#輸入廠商
        word.Selection.Find.Execute("以總計NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("下訂單隔日起")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[12])#輸入工作天數
        word.Selection.Find.Execute("會課:")
        word.Selection.MoveLeft()#前往單價
        word.Selection.MoveRight()#前往單價
        word.Selection.TypeText(depart(v1[6]))#輸入XX課

        #提供複製貼上
        word.Selection.MoveRight(1,2)
        word.Selection.TypeText(v1[23])#

        pass
    elif lastCost>=150000 and lastCost<500000:
                   
        # 複製文件
        #src_file = r"D:\Users\E941\Desktop\採購表單\07_驗收表單\F-AD-2-09-05A維修驗收單.docx"
        
        src_file = r"../../採購表單/01比價會議紀錄表_空白/三家以上-- 比（議）價會議紀錄  金額15萬以上.docx"

        #這段程式碼，在未來想要顯示品名作為word檔案名稱時候，要使用，
        new_filename ="_"#v1[3]
        # 檢查檔案名稱是否合法，若不合法則替換特殊符號
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|',' ']
        for char in invalid_chars:
            if char in new_filename:
                new_filename = new_filename.replace(char, '_')
        dst_file_name = num_value + new_filename + "_15Wto50W議價紀錄"

        #更換目前位置change directory
        # print("目前位置為"+os.getcwd())
        os.chdir("..\\..\\採購表單/")
        dst_folder = str(os.getcwd()) +"\\"
        #print("2目前位置為"+os.getcwd())
            #os.path.join()是把字串黏接起來
        dst_file = os.path.join(dst_folder, dst_file_name + ".docx")
        print("dst_file是"+ dst_file)

        #要將相對位置跳回去
        os.chdir("..\\pyT\\priceRecord")
        #卡在這邊無法copy
        shutil.copy(src_file,dst_file)#這一行srcfile可以相對位置，dst_file只能絕對位置。
        #os.system(f"copy {src_file} {dst_file}")

        # 打開文件
        word = win32.gencache.EnsureDispatch('Word.Application')
        doc = word.Documents.Open(dst_file)

        # 移動游標並進行文字操作
        word.Selection.MoveDown()
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[3])#品名

        word.Selection.Find.Execute("時間：")
        word.Selection.MoveRight()
        today = to_ROC_date_today()
        word.Selection.TypeText(today)
        #for i in range(8):
            #word.Selection.Delete()
        word.Selection.Find.Execute("本案預算NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[17]))#輸入預算
        word.Selection.Find.Execute("廠商名稱")
        word.Selection.MoveRight()
        word.Selection.MoveDown()
        word.Selection.TypeText(v1[16])#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD"+addComma(v1[18]))#輸入最後議價
        
        #提供複製貼上
        word.Selection.Find.Execute("無法再減")
        word.Selection.MoveRight(1,2)
        word.Selection.TypeText(v1[23])#
        word.Selection.MoveRight()
        word.Selection.TypeText("三家議價後: \n"+v1[24])#

        word.Selection.Find.Execute("最後「")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[16])#輸入廠商
        word.Selection.Find.Execute("以總計NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("下訂單隔日起")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[12])#輸入工作天數
        word.Selection.Find.Execute("會課:")
        word.Selection.MoveLeft()#前往單價
        word.Selection.MoveRight()#前往單價
        word.Selection.TypeText(depart(v1[6]))#輸入XX課



        pass
    else:
        print("金額有錯誤，請檢查一下")
        return


    # 儲存文件
    doc.Save()

    # 關閉文件
    doc.Close()

    # 關閉Word應用程序
    word.Quit()
    lblString =  "已成功創建"+dst_file_name+".docx\n\t目前仍未支援自動輸入統一編號，再請自行上網查詢，感謝。\n" + lblString
    label2['text'] = lblString

    #最後資料清空
    v1 = []

def hasBought():
    global num
    global v1
    global lblString

    #取得num的值
    num_value = num.get()
    
        ##這是將excel檔案複製到txt檔案中


    # 指定 Excel 檔案路徑
    #path = r"D:\Users\E941\Desktop\手工驗收單\112工程發包俊龍版2.xlsx"
    path = r"../../112備品俊龍板_0410.xlsx"


    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(path)

    # 宣告欲尋找的字串
    #num = "20231185"

    # 宣告一個空的 List 變數
    #v1 = []

    # 尋找 num 的位置
    for sheet in wb.worksheets:
        for row in sheet.rows:
            for cell in row:
                if str(cell.value) == num_value:
                    for i in range(26):
                        v1.append(str(row[i].value))
                        print(v1[i])


    if len(v1)==0:
        lblString = "於找不到這一筆資料。\n"
        label2['text'] = lblString
        return
    else:
        for i in range(23):
            if v1[i]==None:
                v1[i]="0"

    lastCost =int(v1[18]) 
    if lastCost<8000:

        # 複製文件
        #src_file = r"D:\Users\E941\Desktop\採購表單\07_驗收表單\F-AD-2-09-05A維修驗收單.docx"
        
        src_file = r"../../採購表單/01比價會議紀錄表_空白/已購買_比（議）價會議紀錄金額8000以下.docx"

        #這段程式碼，在未來想要顯示品名作為word檔案名稱時候，要使用，
        new_filename ="_"#v1[3]
        # 檢查檔案名稱是否合法，若不合法則替換特殊符號
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|',' ']
        for char in invalid_chars:
            if char in new_filename:
                new_filename = new_filename.replace(char, '_')
        dst_file_name = num_value + new_filename + "_已購買議價紀錄"

        #更換目前位置change directory
        # print("目前位置為"+os.getcwd())
        os.chdir("..\\..\\採購表單/")
        dst_folder = str(os.getcwd()) +"\\"
        #print("2目前位置為"+os.getcwd())
            #os.path.join()是把字串黏接起來
        dst_file = os.path.join(dst_folder, dst_file_name + ".docx")
        print("dst_file是"+ dst_file)

        #要將相對位置跳回去
        os.chdir("..\\pyT\\priceRecord")
        #卡在這邊無法copy
        shutil.copy(src_file,dst_file)#這一行srcfile可以相對位置，dst_file只能絕對位置。
        #os.system(f"copy {src_file} {dst_file}")

        # 打開文件
        word = win32.gencache.EnsureDispatch('Word.Application')
        doc = word.Documents.Open(dst_file)

        # 移動游標並進行文字操作
        word.Selection.MoveDown()
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[3])#品名

        word.Selection.Find.Execute("時間：")
        word.Selection.MoveRight()
        today = to_ROC_date_today()#可能要去掉民國
        word.Selection.TypeText(today)
        #for i in range(8):
            #word.Selection.Delete()
        word.Selection.Find.Execute("本案預算NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[17]))#輸入預算
        word.Selection.Find.Execute("廠商名稱")
        word.Selection.MoveRight()
        word.Selection.MoveDown()
        word.Selection.TypeText(v1[16])#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD" + addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("本案由申請單位向「")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[16])#輸入廠商
        word.Selection.Find.Execute("最後價格為NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("交貨期限：")
        word.Selection.MoveRight()
        #word.Selection.MoveDown(1,4)

        word.Selection.TypeText(v1[12])#輸入工作天數
        word.Selection.Find.Execute("會課:")
        word.Selection.MoveLeft()
        word.Selection.MoveRight()
        word.Selection.TypeText(depart(v1[6]))#輸入XX課
    else:
        print("金額有錯誤，請檢查一下")
        return


    # 儲存文件
    doc.Save()

    # 關閉文件
    doc.Close()

    # 關閉Word應用程序
    word.Quit()
    lblString =  "已成功創建"+dst_file_name+".docx\n\t目前仍未支援自動輸入統一編號，再請自行上網查詢，感謝。\n" + lblString
    label2['text'] = lblString

    #最後資料清空
    v1 = []

def unicBought():
    global num
    global v1
    global lblString

    #取得num的值
    num_value = num.get()
    
        ##這是將excel檔案複製到txt檔案中


    # 指定 Excel 檔案路徑
    #path = r"D:\Users\E941\Desktop\手工驗收單\112工程發包俊龍版2.xlsx"
    path = r"../../112備品俊龍板_0410.xlsx"


    # 打開 Excel 檔案
    wb = openpyxl.load_workbook(path)

    # 宣告欲尋找的字串
    #num = "20231185"

    # 宣告一個空的 List 變數
    #v1 = []

    # 尋找 num 的位置
    for sheet in wb.worksheets:
        for row in sheet.rows:
            for cell in row:
                if str(cell.value) == num_value:
                    for i in range(26):
                        v1.append(str(row[i].value))
                        print(v1[i])


    if len(v1)==0:
        lblString = "於找不到這一筆資料。\n"
        label2['text'] = lblString
        return
    else:
        for i in range(23):
            if v1[i]==None:
                v1[i]="0"
    lastCost =int(v1[18])
   
    if lastCost>99999:

        # 複製文件
        #src_file = r"D:\Users\E941\Desktop\採購表單\07_驗收表單\F-AD-2-09-05A維修驗收單.docx"
        
        src_file = r"../../採購表單/01比價會議紀錄表_空白/獨家代理-- 比（議）價會議紀錄  金額10萬以上.docx"

        #這段程式碼，在未來想要顯示品名作為word檔案名稱時候，要使用，
        new_filename ="_"#v1[3]
        # 檢查檔案名稱是否合法，若不合法則替換特殊符號
        invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|',' ']
        for char in invalid_chars:
            if char in new_filename:
                new_filename = new_filename.replace(char, '_')
        dst_file_name = num_value + new_filename + "_獨家議價紀錄"

        #更換目前位置change directory
        # print("目前位置為"+os.getcwd())
        os.chdir("..\\..\\採購表單/")
        dst_folder = str(os.getcwd()) +"\\"
        #print("2目前位置為"+os.getcwd())
            #os.path.join()是把字串黏接起來
        dst_file = os.path.join(dst_folder, dst_file_name + ".docx")
        print("dst_file是"+ dst_file)

        #要將相對位置跳回去
        os.chdir("..\\pyT\\priceRecord")
        #卡在這邊無法copy
        shutil.copy(src_file,dst_file)#這一行srcfile可以相對位置，dst_file只能絕對位置。
        #os.system(f"copy {src_file} {dst_file}")

        # 打開文件
        word = win32.gencache.EnsureDispatch('Word.Application')
        doc = word.Documents.Open(dst_file)

        # 移動游標並進行文字操作
        word.Selection.MoveDown()
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[3])#品名

        word.Selection.Find.Execute("時間：")
        word.Selection.MoveRight()
        today = to_ROC_date_today()#可能要去掉民國
        word.Selection.TypeText(today)
        #for i in range(8):
            #word.Selection.Delete()
        word.Selection.Find.Execute("本案預算NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[17]))#輸入預算
        word.Selection.Find.Execute("廠商名稱")
        word.Selection.MoveRight()
        word.Selection.MoveDown()
        word.Selection.TypeText(v1[16])#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD" + addComma(v1[26]))#輸入廠商名稱
        word.Selection.MoveRight()
        word.Selection.TypeText("NTD" + addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("最後「")
        word.Selection.MoveRight()
        word.Selection.TypeText(v1[16])#輸入廠商
        word.Selection.Find.Execute("以總計NTD")
        word.Selection.MoveRight()
        word.Selection.TypeText(addComma(v1[18]))#輸入最後議價
        word.Selection.Find.Execute("交貨期限：下訂單隔日起")
        word.Selection.MoveRight()
        #word.Selection.MoveDown(1,4)

        word.Selection.TypeText(v1[12])#輸入工作天數
        word.Selection.Find.Execute("會課:")
        word.Selection.MoveLeft()
        word.Selection.MoveRight()
        word.Selection.TypeText(depart(v1[6]))#輸入XX課
    else:
        print("金額低於10萬不必獨家議價，請檢查一下")
        return


    # 儲存文件
    doc.Save()

    # 關閉文件
    doc.Close()

    # 關閉Word應用程序
    word.Quit()
    lblString =  "已成功創建"+dst_file_name+".docx\n\t目前仍未支援自動輸入統一編號，再請自行上網查詢，感謝。\n" + lblString
    label2['text'] = lblString

    #最後資料清空
    v1 = []

    pass
btn1 = tk.Button(root, text="正常購買",command=execute)
btn2 = tk.Button(root, text="已購買",command=hasBought)
btn3 = tk.Button(root, text="獨家購買",command=unicBought)

btn1.grid(row=0, column=2)
btn2.grid(row=1, column=0)
btn3.grid(row=1, column=1)

root.mainloop()
